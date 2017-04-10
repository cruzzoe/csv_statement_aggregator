"""Script to aggregate CSV statements into an XLSX document.
    Can be used to take 6 months of statements into one document for easier
    reviewing of spending patterns.
"""


import csv
import datetime
from openpyxl import Workbook
import os


# Enter directory where CSV statements will be saved.
DIRECTORY = ''

list_of_files = {}


def walkDirectory(directory):
    """Find all csv files in the supplied directory."""
    for (dirpath, dirnames, filenames) in os.walk(directory):
        for filename in filenames:
            if filename.endswith('.csv'):
                list_of_files[filename] = os.sep.join([dirpath, filename])

    return list_of_files


def createWorkbook():
    """Create xlsx workbook."""
    wb = Workbook()
    ws = wb.active
    return ws, wb


def unpackCsvFiles(list_of_files):
    """Take Csv filess and load as a dictionary."""
    aggregate = []

    for filePath in list_of_files.values():
        csvFile = open(filePath)
        reader = csv.DictReader(csvFile)

        for row in reader:
            aggregate.append(row)
    return aggregate


def castDateStrToDate(aggregate):
    """Convert dates to date objects."""
    for row in aggregate:
        dateStr = row['Transaction Date']
        transactionDate = datetime.datetime.strptime(dateStr, '%d/%m/%Y').date()
        row['Date'] = transactionDate
    return aggregate


def updateXlsx(aggregate, ws):
    """Update Xlsx with data obtained from Csv."""
    aggregate.sort(key=lambda x: x['Date'])

    for i, row in enumerate(aggregate):
        # Exclude Thank you for payment rows from statements
        if 'THANK YOU' in row['Description']:
            continue

        i += 1

        for j, col in enumerate(row.values()):
            j += 1
            _ = ws.cell(column=j, row=i, value="{0}".format(col))


def saveXlsx(wb):
    """Save the Xlsx file."""
    wb.save("statement_agg.xlsx")


def run():
    csvFiles = walkDirectory(DIRECTORY)
    wb, ws = createWorkbook()
    agg = unpackCsvFiles(csvFiles)
    agg = castDateStrToDate(agg)
    updateXlsx(agg, ws)
    saveXlsx(wb)
